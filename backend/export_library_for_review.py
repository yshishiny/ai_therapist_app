"""
export_library_for_review.py — the content library as a review worksheet
--------------------------------------------------------------------------
Every resource in one sheet so a clinician can check the citations and record
a decision against each.

Why this exists: 33 of the 91 items were never citation-checked. Their
collector wrote the file and its audit never ran, so nobody has confirmed the
titles, authors or claims. They are marked in the app, but marking them is not
the same as checking them.

Checking is awkward in the app because only videos carry a URL -- books and
papers have none, so verifying one means retyping its title into a search
engine. This sheet builds that search link for every row, clearly labelled as
a SEARCH and not as the source itself. A fabricated DOI would be worse than no
link at all.

Usage:
    DATABASE_HOST=... DB_PASSWORD=... python backend/export_library_for_review.py [out.xlsx]
"""

from __future__ import annotations

import asyncio
import json
import os
import pathlib
import sys
import urllib.parse

import asyncpg
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

FONT = "Arial"
INK = "1F2937"
MUTED = "6B7280"
HEAD_FILL = PatternFill("solid", fgColor="5262AD")
WARN_FILL = PatternFill("solid", fgColor="FDECEC")   # unchecked rows
OK_FILL = PatternFill("solid", fgColor="F3F7F2")     # already checked
DECIDE_FILL = PatternFill("solid", fgColor="FFF4CE") # for the reviewer to fill
thin = Side(style="thin", color="D5D8E0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

TOPIC_LABEL = {
    "addiction": "Addiction",
    "anxiety": "Anxiety",
    "emotional_intelligence": "Emotional Intelligence",
    "depression": "Depression",
    "ptsd": "PTSD",
    "personality_disorders": "Personality Disorders",
}

COLUMNS = [
    ("#", 5),
    ("Type", 9),
    ("Topic", 20),
    ("Citations checked?", 17),
    ("Title", 60),
    ("Author / source", 30),
    ("Link (video) or search", 46),
    ("Why it is in the library", 70),
    ("Your decision", 16),
    ("Notes", 34),
]

DECISIONS = ["Keep", "Remove", "Needs correction", "Unsure - discuss"]


def lookup_url(category: str, title: str, author: str | None, file_url: str | None) -> tuple[str, str]:
    """(url, what the url actually is). Never presents a search as a source."""
    if file_url:
        return file_url, "the item itself"
    query = " ".join(filter(None, [title, author]))
    q = urllib.parse.quote_plus(query[:250])
    if category == "PAPER":
        return f"https://scholar.google.com/scholar?q={q}", "a search — verify the match"
    if category == "BOOK":
        return f"https://www.google.com/search?tbm=bks&q={q}", "a search — verify the match"
    return f"https://www.google.com/search?q={q}", "a search — verify the match"


async def fetch_rows(conn) -> list[dict]:
    rows = await conn.fetch(
        """
        SELECT title, author, category, description, file_url, tags
        FROM resources
        ORDER BY (tags ? 'unverified') DESC, category, title
        """
    )
    out = []
    for r in rows:
        tags = r["tags"]
        tags = json.loads(tags) if isinstance(tags, str) else (tags or [])
        topic = next((t for t in tags if t in TOPIC_LABEL), None)
        out.append(
            {
                "title": r["title"],
                "author": r["author"],
                "category": r["category"],
                "description": r["description"],
                "file_url": r["file_url"],
                "topic": TOPIC_LABEL.get(topic, topic or "—"),
                "verified": "verified" in tags,
            }
        )
    return out


def build(rows: list[dict], out: pathlib.Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Library review"
    ws.freeze_panes = "A3"

    unchecked = sum(1 for r in rows if not r["verified"])
    ws["A1"] = (
        f"{len(rows)} items in the content library. {unchecked} have NOT been citation-checked "
        f"(shaded red) — nobody has confirmed those titles, authors or claims. "
        "Open the link, confirm the item is real and described accurately, then record a decision. "
        "Links marked \"a search\" are a search, not the source: check the result actually matches."
    )
    ws["A1"].font = Font(name=FONT, size=10.5, color=INK)
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    ws.row_dimensions[1].height = 46

    for j, (label, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(j)].width = width
        c = ws.cell(row=2, column=j, value=label)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = HEAD_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[2].height = 30

    for i, r in enumerate(rows, start=1):
        row = i + 2
        url, kind = lookup_url(r["category"], r["title"], r["author"], r["file_url"])
        values = [
            i,
            r["category"].title(),
            r["topic"],
            "Checked" if r["verified"] else "NOT CHECKED",
            r["title"],
            r["author"] or "—",
            url,
            (r["description"] or "").strip(),
            "",
            "",
        ]
        for j, v in enumerate(values, start=1):
            c = ws.cell(row=row, column=j, value=v)
            c.font = Font(name=FONT, size=9.5, color=INK)
            c.alignment = Alignment(wrap_text=j in (5, 6, 8), vertical="top")
            c.border = BORDER
            if not r["verified"]:
                c.fill = WARN_FILL
            elif j <= 4:
                c.fill = OK_FILL
            if j == 9:
                c.fill = DECIDE_FILL
        # Make the link clickable, and say underneath what it actually points at.
        link = ws.cell(row=row, column=7)
        link.hyperlink = url
        link.value = kind
        link.font = Font(name=FONT, size=9.5, color="0563C1", underline="single")
        ws.row_dimensions[row].height = 42

    dv = DataValidation(type="list", formula1='"' + ",".join(DECISIONS) + '"', allow_blank=True)
    dv.promptTitle = "Your decision"
    dv.prompt = "Keep / Remove / Needs correction / Unsure"
    ws.add_data_validation(dv)
    dv.add(f"I3:I{len(rows) + 2}")

    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}{len(rows) + 2}"

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"wrote {out}  ({len(rows)} items, {unchecked} unchecked)")


async def main() -> None:
    out = pathlib.Path(sys.argv[1]) if len(sys.argv) > 1 else pathlib.Path(
        "docs/forms/content_library_review.xlsx"
    )
    conn = await asyncpg.connect(
        user=os.getenv("DB_USER", "appuser"),
        password=os.environ["DB_PASSWORD"],
        database=os.getenv("DB_NAME", "ai_therapist"),
        host=os.environ["DATABASE_HOST"],
        ssl="require",
    )
    rows = await fetch_rows(conn)
    await conn.close()
    build(rows, out)


asyncio.run(main())
