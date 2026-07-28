import csv
import io
import json
from datetime import date, datetime

from fastapi import APIRouter, Depends, File, HTTPException, Query, Response, UploadFile, status

from backend.core.dependencies_access import require_permission
from backend.src.core.dependencies import (
    RequestContext,
    get_clinician_context,
    get_clinician_service,
    get_patient_service,
)
from backend.src.schemas.patients import (
    PatientBulkUploadSummaryOut,
    PatientCreateIn,
    PatientListOut,
    PatientOut,
    PatientUpdateIn,
)
from backend.src.services.clinician_service_db import ClinicianServiceDb
from backend.src.services.patient_service_db import PatientServiceDb

router = APIRouter(prefix='/patients', tags=['patients'])

BULK_TEMPLATE_COLUMNS = [
    'full_name', 'gender', 'dob', 'diagnosis', 'risk', 'status', 'phone', 'email', 'therapist_email',
]
BULK_UPLOAD_MAX_ROWS = 200


@router.get('', response_model=PatientListOut)
async def list_patients(
    limit: int = Query(50, ge=1, le=200, description='Page size.'),
    offset: int = Query(0, ge=0, description='Rows to skip.'),
    mine: bool = Query(False, description="Scope to the calling clinician's own caseload."),
    context: RequestContext = Depends(get_clinician_context),
    _perm=require_permission('patients.view'),
    service: PatientServiceDb = Depends(get_patient_service),
):
    """One page of the organisation's patients, with the matching total.

    RESPONSE SHAPE CHANGE: this returns `{items, total, limit, offset}`, not a
    bare array. The endpoint has always been paged -- `limit` defaulted to 50
    and silently truncated -- but returned no total, so a client receiving 50
    patients could not tell whether that was the whole caseload or the first
    page of nine. Every caller was updated in the same change.

    GET /patients/{id} is unaffected and still returns a single object.
    """
    therapist_id = context.user_id if mine else None
    return await service.list_patients(
        org_id=context.org_id, limit=limit, offset=offset, therapist_id=therapist_id
    )


# NOTE: the bulk routes are declared BEFORE GET '/{patient_id}' on purpose --
# otherwise FastAPI would match 'bulk-template' as a patient_id.
@router.get('/bulk-template')
async def download_bulk_template(
    context: RequestContext = Depends(get_clinician_context),
    _perm=require_permission('patients.manage'),
):
    # Header only, no sample rows. This template used to ship two "Example
    # Patient" rows, which the importer cannot distinguish from real ones -- so
    # downloading it and uploading it back unchanged created two fictional
    # patients in a clinical system. Worked examples live on the Instructions
    # sheet of the intake workbook instead, where they cannot be imported.
    buffer = io.StringIO()
    writer = csv.writer(buffer, lineterminator='\r\n')
    writer.writerow(BULK_TEMPLATE_COLUMNS)
    return Response(
        content=buffer.getvalue(),
        media_type='text/csv',
        headers={'Content-Disposition': 'attachment; filename="patients_template.csv"'},
    )


def _clean_cell(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _normalize_row(raw_row: dict) -> dict:
    return {
        str(key).strip().lower(): _clean_cell(value)
        for key, value in raw_row.items()
        if key is not None
    }


def _parse_xlsx_rows(raw: bytes) -> list[dict]:
    """Read the 'Patients' sheet of the intake workbook into row dicts.

    Deliberately reads only the sheet named 'Patients' when one exists, so the
    Instructions sheet -- which carries worked examples -- can never be mistaken
    for data. Dates are stringified back to YYYY-MM-DD: Excel silently converts
    a typed date into a datetime, and the row validator expects the text form.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency is declared
        raise HTTPException(
            status_code=400,
            detail='Spreadsheet upload is unavailable on this server; please upload a .csv instead.',
        ) from exc

    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f'The spreadsheet could not be read: {exc}') from exc

    ws = wb['Patients'] if 'Patients' in wb.sheetnames else wb[wb.sheetnames[0]]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise HTTPException(status_code=400, detail='The spreadsheet is empty.') from None

    header = [str(h).strip().lower() if h is not None else '' for h in header_row]
    if 'full_name' not in header:
        raise HTTPException(
            status_code=400,
            detail="The spreadsheet's first row is not the template header (no 'full_name' column). "
                   "Use the 'Patients' sheet of the intake form without renaming its columns.",
        )

    # The form pre-fills the assigned clinician down the whole sheet, so a row
    # carrying only that value is still an unused template row. Judge emptiness
    # on everything EXCEPT that column -- but no further: a row with a diagnosis
    # or a date and no name is a real mistake and must be reported, not skipped.
    significant = [i for i, key in enumerate(header) if key != 'therapist_email']

    out: list[dict] = []
    # Sheet row numbers, carried through so an error can name the line the
    # clinician is actually looking at. Counting positions in the parsed list
    # instead would drift by one for every blank row skipped above it, and send
    # someone to correct the wrong patient.
    sheet_row = 1  # the header we just consumed
    for values in rows_iter:
        sheet_row += 1
        if values is None:
            continue
        if all(
            i >= len(values) or values[i] is None or str(values[i]).strip() == ''
            for i in significant
        ):
            continue  # an unused template row, not a patient
        row: dict = {}
        for key, value in zip(header, values):
            if not key:
                continue
            if value is None:
                row[key] = None
            elif isinstance(value, datetime):
                row[key] = value.date().isoformat()
            elif isinstance(value, date):
                row[key] = value.isoformat()
            else:
                row[key] = str(value).strip()
        row['__sheet_row__'] = sheet_row
        out.append(row)
    return out


def _parse_bulk_rows(filename: str, content_type: str | None, raw: bytes) -> list:
    """Return the list of raw data rows (dicts for CSV; anything for JSON --
    non-dict entries become per-row errors later). Raises HTTPException for
    file-level problems."""
    name = (filename or '').lower()
    is_csv = name.endswith('.csv') or content_type in ('text/csv', 'application/csv')
    is_json = name.endswith('.json') or content_type in ('application/json', 'text/json')
    is_xlsx = name.endswith(('.xlsx', '.xlsm')) or content_type in (
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'application/vnd.ms-excel.sheet.macroEnabled.12',
    )
    if not is_csv and not is_json and not is_xlsx:
        raise HTTPException(status_code=400, detail='Upload a .xlsx, .csv or .json file.')

    # The intake form we send clinicians is a workbook, so accept it directly
    # rather than making them re-save it as CSV -- a conversion step is one more
    # place for a column to be dropped or a date to be silently reformatted.
    if is_xlsx:
        return _parse_xlsx_rows(raw)

    try:
        text = raw.decode('utf-8-sig')
    except UnicodeDecodeError as exc:
        raise HTTPException(status_code=400, detail='The file is not valid UTF-8 text.') from exc

    if is_csv:
        reader = csv.DictReader(io.StringIO(text))
        if reader.fieldnames is None:
            raise HTTPException(status_code=400, detail='The CSV file is empty.')
        header = [str(f).strip().lower() for f in reader.fieldnames if f]
        if 'full_name' not in header:
            raise HTTPException(
                status_code=400,
                detail="The CSV header does not match the template (missing 'full_name' column). "
                       'Download the template from GET /patients/bulk-template.',
            )
        return list(reader)

    try:
        payload = json.loads(text)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail=f'Not valid JSON: {exc}') from exc
    if not isinstance(payload, list):
        raise HTTPException(
            status_code=400,
            detail='The JSON file must contain an array of patient objects.',
        )
    return payload


@router.post('/bulk-upload', response_model=PatientBulkUploadSummaryOut)
async def bulk_upload_patients(
    file: UploadFile = File(...),
    context: RequestContext = Depends(get_clinician_context),
    _perm=require_permission('patients.manage'),
    service: PatientServiceDb = Depends(get_patient_service),
):
    raw = await file.read()
    rows = _parse_bulk_rows(file.filename or '', file.content_type, raw)

    if not rows:
        raise HTTPException(status_code=400, detail='No data rows found in the file.')
    if len(rows) > BULK_UPLOAD_MAX_ROWS:
        raise HTTPException(
            status_code=400,
            detail=f'Too many rows ({len(rows)}). '
                   f'The limit is {BULK_UPLOAD_MAX_ROWS} patients per upload.',
        )

    created = 0
    errors: list[dict] = []
    email_cache: dict[str, str | None] = {}

    # Row numbers match the source file: header = row 1, first data row = 2.
    # A spreadsheet carries its own true row number, because blank rows are
    # skipped during parsing and a positional count would drift past them.
    for position, raw_row in enumerate(rows, start=2):
        row_number = position
        if isinstance(raw_row, dict) and '__sheet_row__' in raw_row:
            raw_row = dict(raw_row)
            row_number = raw_row.pop('__sheet_row__')
        if not isinstance(raw_row, dict):
            errors.append({'row': row_number, 'error': 'Row must be an object of patient fields.'})
            continue
        row = _normalize_row(raw_row)

        if not row.get('full_name'):
            errors.append({'row': row_number, 'error': 'full_name is required.'})
            continue

        dob = row.get('dob')
        if dob is not None:
            try:
                datetime.strptime(dob, '%Y-%m-%d')
            except ValueError:
                errors.append({'row': row_number, 'error': f"dob '{dob}' must be in YYYY-MM-DD format."})
                continue

        therapist_email = row.get('therapist_email')
        if therapist_email is not None:
            cache_key = therapist_email.lower()
            if cache_key not in email_cache:
                email_cache[cache_key] = await service.find_clinician_id_by_email(
                    email=therapist_email, org_id=context.org_id
                )
            therapist_id = email_cache[cache_key]
            if therapist_id is None:
                errors.append({
                    'row': row_number,
                    'error': f"No clinician with email '{therapist_email}' was found in this practice.",
                })
                continue
        else:
            therapist_id = context.user_id

        payload = {'full_name': row['full_name']}
        for field in ('gender', 'diagnosis', 'risk', 'status', 'phone', 'email', 'dob'):
            value = row.get(field)
            if value is not None:
                payload[field] = value

        try:
            body = PatientCreateIn(**payload)
            await service.create_patient(
                body=body,
                org_id=context.org_id,
                therapist_id=therapist_id,
                source='BULK_UPLOAD',
                created_by=context.user_id,
                source_detail=f'Imported from {file.filename or "an uploaded file"}',
            )
            created += 1
        except Exception as exc:
            errors.append({'row': row_number, 'error': f'Could not create patient: {exc}'})

    return {'created': created, 'failed': len(errors), 'errors': errors}


@router.get('/{patient_id}', response_model=PatientOut)
async def get_patient(
    patient_id: str,
    context: RequestContext = Depends(get_clinician_context),
    _perm=require_permission('patients.view'),
    service: PatientServiceDb = Depends(get_patient_service),
):
    patient = await service.get_patient(patient_id=patient_id, org_id=context.org_id)
    if patient is None:
        raise HTTPException(status_code=404, detail='Patient not found.')
    return patient


@router.patch('/{patient_id}', response_model=PatientOut)
async def update_patient(
    patient_id: str,
    body: PatientUpdateIn,
    context: RequestContext = Depends(get_clinician_context),
    _perm=require_permission('patients.manage'),
    service: PatientServiceDb = Depends(get_patient_service),
    clinician_service: ClinicianServiceDb = Depends(get_clinician_service),
):
    if body.therapist_id is not None:
        valid = await clinician_service.clinician_exists(
            clinician_id=body.therapist_id, org_id=context.org_id
        )
        if not valid:
            raise HTTPException(status_code=400, detail='That clinician was not found in this practice.')
    patient = await service.update_patient(patient_id=patient_id, org_id=context.org_id, body=body)
    if patient is None:
        raise HTTPException(status_code=404, detail='Patient not found.')
    return patient


@router.post('', response_model=PatientOut, status_code=status.HTTP_201_CREATED)
async def create_patient(
    body: PatientCreateIn,
    context: RequestContext = Depends(get_clinician_context),
    _perm=require_permission('patients.manage'),
    service: PatientServiceDb = Depends(get_patient_service),
):
    return await service.create_patient(
        body=body,
        org_id=context.org_id,
        therapist_id=context.user_id,
        source='MANUAL',
        created_by=context.user_id,
        source_detail='Added in the practice console',
    )
