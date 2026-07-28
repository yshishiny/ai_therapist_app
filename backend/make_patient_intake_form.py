"""
make_patient_intake_form.py — build the patient intake spreadsheet
------------------------------------------------------------------
The form a practice sends to a clinician so they can list their real patients,
return it, and have it imported through POST /patients/bulk-upload.

Its columns are exactly BULK_TEMPLATE_COLUMNS from
backend/src/api/routes/patients_db_wired.py. If that list changes, change it
here in the same commit or the import will reject every row.

One deliberate difference from the CSV that endpoint serves: the worked
examples live on the Instructions sheet, NOT in the data grid. Example rows
sitting in the grid are indistinguishable from real ones to the importer, so
anyone who filled the sheet in without deleting them would create fictional
patients in a clinical system.

Usage:  python backend/make_patient_intake_form.py [output.xlsx]
"""

from __future__ import annotations

import pathlib
import sys

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Must match BULK_TEMPLATE_COLUMNS exactly, in order.
COLUMNS = [
    ("full_name", "Full name", True, 30),
    ("gender", "Gender", False, 14),
    ("dob", "Date of birth", False, 15),
    ("diagnosis", "Diagnosis / presenting concern", False, 34),
    ("risk", "Risk", False, 10),
    ("status", "Status", False, 14),
    ("phone", "Phone", False, 18),
    ("email", "Email", False, 28),
    ("therapist_email", "Assigned clinician (email)", False, 30),
]

GENDERS = ["Female", "Male", "Non-binary", "Prefer not to say"]
RISKS = ["Low", "Med", "High"]
STATUSES = ["Active", "Intake", "Maintenance", "Discharged"]

DEFAULT_THERAPIST_EMAIL = "heba.moustafa5@gmail.com"

MAX_ROWS = 200  # BULK_UPLOAD_MAX_ROWS
DATA_ROWS = 120  # pre-formatted blank rows; the cap is enforced server-side

FONT = "Arial"
INK = "1F2937"
MUTED = "6B7280"
ACCENT = "5262AD"
HEAD_FILL = PatternFill("solid", fgColor="5262AD")
REQ_FILL = PatternFill("solid", fgColor="FFF4CE")   # yellow: must be filled in
BAND_FILL = PatternFill("solid", fgColor="F5F6FA")
NOTE_FILL = PatternFill("solid", fgColor="FDECEC")
thin = Side(style="thin", color="D5D8E0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def _title(ws, cell: str, text: str, size: int = 14) -> None:
    ws[cell] = text
    ws[cell].font = Font(name=FONT, size=size, bold=True, color=INK)


def _body(ws, cell: str, text: str, *, bold=False, color=INK, size=10.5, wrap=True) -> None:
    ws[cell] = text
    ws[cell].font = Font(name=FONT, size=size, bold=bold, color=color)
    ws[cell].alignment = Alignment(wrap_text=wrap, vertical="top")


def build_instructions(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 74

    _title(ws, "B2", "Patient list — intake form", 16)
    _body(
        ws, "B3",
        "Please list the patients you would like set up in the AI Therapist system, one per row on "
        "the 'Patients' sheet, then return this file to your practice administrator.",
        color=MUTED,
    )
    ws.merge_cells("B3:D3")
    ws.row_dimensions[3].height = 30

    # Handling notice first: this file will hold identifiable patient data.
    _body(ws, "B5", "Before you begin", bold=True, size=12)
    ws["B6"].fill = NOTE_FILL
    ws.merge_cells("B6:D8")
    _body(
        ws, "B6",
        "This form will contain identifiable patient information once you fill it in. Treat the "
        "completed file as clinical data: return it by a secure route agreed with your practice "
        "(not ordinary email), keep it only until it has been imported, and delete your copy "
        "afterwards. Enter only patients you have a lawful basis to register, and record no more "
        "than the fields asked for — please do not add clinical notes or history to this sheet.",
    )
    ws.row_dimensions[6].height = 46

    _body(ws, "B10", "How to fill it in", bold=True, size=12)
    steps = [
        "Use the 'Patients' sheet. One patient per row, starting at row 2, directly under the headers. Do not rename, reorder or "
        "remove the header row — the importer matches on those exact column names.",
        "Only 'Full name' is required. Every other column may be left blank, and blanks are simply "
        "not recorded rather than guessed at.",
        f"Up to {MAX_ROWS} patients per file. For a larger list, split it across several copies.",
        "Gender, Risk and Status are drop-downs — pick from the list so the values import cleanly.",
        "Dates of birth must be written as YYYY-MM-DD, for example 1990-01-15. Any other format is "
        "rejected, with the row number reported back so it can be corrected.",
    ]
    row = 11
    for i, step in enumerate(steps, 1):
        _body(ws, f"B{row}", f"{i}.", bold=True, wrap=False)
        _body(ws, f"C{row}", step)
        ws.merge_cells(f"C{row}:D{row}")
        ws.row_dimensions[row].height = 30
        row += 1

    row += 1
    _body(ws, f"B{row}", "What each column means", bold=True, size=12)
    row += 1
    header_row = row
    for j, text in enumerate(["Column", "Required", "Notes"]):
        c = ws.cell(row=header_row, column=2 + j, value=text)
        c.font = Font(name=FONT, size=10.5, bold=True, color="FFFFFF")
        c.fill = HEAD_FILL
        c.alignment = Alignment(vertical="center")
    ws.row_dimensions[header_row].height = 20

    notes = {
        "full_name": "The patient's full name as you would write it in their record.",
        "gender": "Choose from the drop-down, or leave blank.",
        "dob": "YYYY-MM-DD. Used to show the patient's age on their chart.",
        "diagnosis": "Working diagnosis or presenting concern, in a few words.",
        "risk": "Your current clinical judgement. Leave blank if not assessed — it defaults to Low, "
                "so set it deliberately rather than relying on the default.",
        "status": "Where the patient is in your service. Defaults to Active.",
        "phone": "Any format. Include the country code if you have it.",
        "email": "Only if the patient is to be given access to the patient app.",
        "therapist_email": f"Pre-filled with {DEFAULT_THERAPIST_EMAIL}, so you can leave it as it "
                           "is. Change it only if a patient belongs to a different clinician, and "
                           "use the address they sign in with. An address the practice does not "
                           "recognise is reported back rather than guessed at.",
    }
    r = header_row + 1
    for key, label, required, _w in COLUMNS:
        ws.cell(row=r, column=2, value=label).font = Font(name=FONT, size=10.5, bold=True, color=INK)
        req = ws.cell(row=r, column=3, value="Required" if required else "Optional")
        req.font = Font(name=FONT, size=10.5, bold=required, color="B45309" if required else MUTED)
        note = ws.cell(row=r, column=4, value=notes[key])
        note.font = Font(name=FONT, size=10.5, color=INK)
        note.alignment = Alignment(wrap_text=True, vertical="top")
        for col in (2, 3, 4):
            ws.cell(row=r, column=col).border = BORDER
            if (r - header_row) % 2 == 0:
                ws.cell(row=r, column=col).fill = BAND_FILL
        ws.row_dimensions[r].height = 30 if key == "therapist_email" else 22
        r += 1

    r += 1
    _body(ws, f"B{r}", "Worked example", bold=True, size=12)
    r += 1
    _body(
        ws, f"C{r}",
        "This is how two completed rows should look. They are shown here rather than in the grid on "
        "purpose: an example left sitting in the 'Patients' sheet would be imported as a real "
        "patient, because nothing distinguishes it from one.",
        color=MUTED,
    )
    ws.merge_cells(f"C{r}:D{r}")
    ws.row_dimensions[r].height = 30
    r += 2

    example_head = r
    for j, (_key, label, _req, _w) in enumerate(COLUMNS):
        c = ws.cell(row=example_head, column=2 + j, value=label)
        c.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
        c.fill = HEAD_FILL
        c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[example_head].height = 26

    examples = [
        ["Layla Hassan", "Female", "1994-03-22", "Generalised anxiety", "Med", "Active",
         "+20 100 555 0134", "layla.hassan@example.com", DEFAULT_THERAPIST_EMAIL],
        ["Omar Naguib", "Male", "1981-11-04", "Adjustment disorder", "Low", "Intake",
         "+20 100 555 0187", "", DEFAULT_THERAPIST_EMAIL],
    ]
    for k, ex in enumerate(examples, start=1):
        for j, val in enumerate(ex):
            c = ws.cell(row=example_head + k, column=2 + j, value=val)
            c.font = Font(name=FONT, size=9, color=INK)
            c.border = BORDER
            c.alignment = Alignment(vertical="center")
    # The example grid is wider than the notes table; widen only what it needs.
    for j, (_key, _label, _req, width) in enumerate(COLUMNS):
        letter = get_column_letter(2 + j)
        if j >= 3:  # leave B/C/D sized for the prose above
            ws.column_dimensions[letter].width = max(ws.column_dimensions[letter].width or 0, 18)


def build_patients(ws) -> None:
    ws.freeze_panes = "A2"

    for j, (key, label, required, width) in enumerate(COLUMNS, start=1):
        letter = get_column_letter(j)
        ws.column_dimensions[letter].width = width
        c = ws.cell(row=1, column=j, value=key)  # the importer matches on the KEY
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = HEAD_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        # The plain-English name lives in a comment, NOT in a second header row.
        # A second header row is a trap: saved as CSV it parses as data and
        # creates a patient literally named "Full name *".
        requirement = "required" if required else "optional"
        c.comment = Comment(
            f"{label} ({requirement})\n\n"
            "Do not rename this cell - the importer matches on it.",
            "AI Therapist",
        )
    ws.row_dimensions[1].height = 24

    for r in range(2, 2 + DATA_ROWS):
        for j, (key, _label, required, _w) in enumerate(COLUMNS, start=1):
            c = ws.cell(row=r, column=j)
            c.font = Font(name=FONT, size=10, color=INK)
            c.border = BORDER
            if required:
                c.fill = REQ_FILL
            if key == "dob":
                c.number_format = "@"  # text, so 1990-01-15 is never re-formatted as a date
            if key == "therapist_email":
                # Pre-filled so nobody has to type or misspell it. The importer
                # ignores this column when deciding whether a row is an unused
                # template row, so a sheet of pre-filled blanks stays blank.
                c.value = DEFAULT_THERAPIST_EMAIL
                c.font = Font(name=FONT, size=10, color=MUTED)

    def add_list(col_key: str, values: list[str], prompt: str) -> None:
        idx = next(i for i, (k, *_r) in enumerate(COLUMNS, start=1) if k == col_key)
        letter = get_column_letter(idx)
        dv = DataValidation(
            type="list",
            formula1='"' + ",".join(values) + '"',
            allow_blank=True,
            showDropDown=False,
        )
        dv.promptTitle = prompt
        dv.prompt = "Choose one, or leave blank."
        dv.errorTitle = "Not one of the accepted values"
        dv.error = "Pick from the drop-down: " + ", ".join(values)
        ws.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}{1 + DATA_ROWS}")

    add_list("gender", GENDERS, "Gender")
    add_list("risk", RISKS, "Risk level")
    add_list("status", STATUSES, "Status")

    # NO footer note on this sheet, deliberately. Anything written into a cell
    # below the grid is read back as a patient row -- an earlier draft put the
    # guidance here and it imported as a patient named "Yellow columns are
    # required...". All guidance lives on the Instructions sheet and in the
    # header comments, where the importer cannot see it.


def main() -> None:
    out = pathlib.Path(sys.argv[1]) if len(sys.argv) > 1 else pathlib.Path(
        "docs/forms/patient_intake_form.xlsx"
    )
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    build_instructions(wb.active)
    wb.active.title = "Instructions"
    build_patients(wb.create_sheet("Patients"))
    wb.save(out)
    print(f"wrote {out}")


if __name__ == "__main__":
    main()
